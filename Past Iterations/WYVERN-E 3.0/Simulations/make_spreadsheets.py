#!/usr/bin/env python3
"""Build formatted Excel workbooks from the WYVERN-E 2.0 simulation CSVs.

Produces:
  Flight_Atmospherics/WYVERN_E2_atmospherics.xlsx  (Summary, Flight State, ISA Reference)
  CFD/WYVERN_E2_airfoil_polars.xlsx                (Summary, Polars)

Raw data rows are the computed scientific values; SUMMARY cells use live Excel
formulas (MAX/AVERAGE/SLOPE/INDEX-MATCH) so the books recompute if data changes.
Run the CSV-producing scripts first (expand_flight.py, run_airfoil_cfd.py).
"""
import os, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FA   = os.path.join(HERE, "Flight_Atmospherics")
CFD  = os.path.join(HERE, "CFD")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BASE     = Font(name="Arial", size=10)
TITLE    = Font(name="Arial", bold=True, size=13, color="1F3864")
KEYFILL  = PatternFill("solid", fgColor="FFF2CC")
THIN     = Side(style="thin", color="D9D9D9")
BORDER   = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

def read_csv(path):
    with open(path) as f:
        r=csv.reader(f); rows=list(r)
    return rows[0], rows[1:]

def num(x):
    try: return float(x)
    except (ValueError, TypeError): return x

def data_sheet(wb, name, header, rows, numfmts=None, freeze="A2"):
    ws=wb.create_sheet(name)
    ws.append(header)
    for c in range(1,len(header)+1):
        cell=ws.cell(1,c); cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for row in rows:
        ws.append([num(v) for v in row])
    # formatting
    for r in range(2, ws.max_row+1):
        for c in range(1, len(header)+1):
            cell=ws.cell(r,c); cell.font=BASE; cell.border=BORDER
            if numfmts and header[c-1] in numfmts and isinstance(cell.value,(int,float)):
                cell.number_format=numfmts[header[c-1]]
    for c in range(1,len(header)+1):
        ws.column_dimensions[get_column_letter(c)].width=max(10, min(16, len(str(header[c-1]))+3))
    ws.freeze_panes=freeze
    return ws

def title_row(ws, text, span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1); c.value=text; c.font=TITLE

def kv(ws, r, label, formula, fmt=None, unit=""):
    ws.cell(r,1,label).font=Font(name="Arial",bold=True,size=10)
    cell=ws.cell(r,2,formula); cell.font=BASE; cell.fill=KEYFILL; cell.border=BORDER
    if fmt: cell.number_format=fmt
    if unit: ws.cell(r,3,unit).font=Font(name="Arial",italic=True,size=9,color="808080")

# ----------------------------------------------------------------------------- atmospherics
def build_atmospherics():
    fh,fr = read_csv(os.path.join(FA,"flight_state.csv"))
    ih,ir = read_csv(os.path.join(FA,"isa_reference.csv"))
    wb=Workbook(); wb.remove(wb.active)
    # Summary sheet (formulas reference the data sheets)
    s=wb.create_sheet("Summary"); title_row(s,"WYVERN-E 2.0 — Flight Atmospherics (PDR-005)",3)
    n=len(fr)+1
    rows=[("Apogee", f"=MAX('Flight State'!C2:C{n})","0.0"," m"),
          ("Max velocity", f"=MAX('Flight State'!D2:D{n})","0.0"," m/s"),
          ("Max Mach", f"=MAX('Flight State'!E2:E{n})","0.000"," —"),
          ("Max dynamic pressure q", f"=MAX('Flight State'!F2:F{n})","0"," Pa"),
          ("Max acceleration", f"=MAX('Flight State'!L2:L{n})","0.0"," g"),
          ("Max fin Reynolds", f"=MAX('Flight State'!Y2:Y{n})","0"," —"),
          ("Max body Reynolds", f"=MAX('Flight State'!X2:X{n})","0"," —"),
          ("Time to apogee", f"=MAX('Flight State'!A2:A{n})","0.0"," s"),
          ("Sea-level density (ISA)", "=INDEX('ISA Reference'!H2:H200,1)","0.000"," kg/m³"),
          ("Sea-level speed of sound","=INDEX('ISA Reference'!I2:I200,1)","0.0"," m/s")]
    for i,(lab,f,fmt,unit) in enumerate(rows): kv(s,i+3,lab,f,fmt,unit)
    s.column_dimensions["A"].width=26; s.column_dimensions["B"].width=14; s.column_dimensions["C"].width=8
    s.cell(15,1,"Trajectory engine: ../run_sims.py combined 2-stage (G78→F25); atmosphere: U.S. Std 1976.").font=Font(name="Arial",italic=True,size=9,color="808080")
    # data sheets
    ffmt={"t":"0.00","h":"0.0","v":"0.0","mach":"0.000","q":"0","drag":"0.00","thrust":"0.0",
          "weight":"0.00","mass":"0.000","accel":"0.0","accel_g":"0.00","TW":"0.00","T":"0.00",
          "T_C":"0.0","p":"0","p_kPa":"0.00","rho":"0.0000","a_snd":"0.0","mu":"0.00E+00",
          "nu":"0.00E+00","sigma":"0.0000","delta":"0.0000","Re_body":"0","Re_fin":"0"}
    data_sheet(wb,"Flight State",fh,fr,ffmt)
    ifmt={"z_m":"0","h_geopot_m":"0.0","T_K":"0.00","T_C":"0.0","p_Pa":"0","p_kPa":"0.00",
          "p_atm":"0.0000","rho_kgm3":"0.0000","a_ms":"0.0","mu_Pas":"0.00E+00","nu_m2s":"0.00E+00",
          "sigma":"0.0000","delta":"0.0000","theta":"0.0000"}
    data_sheet(wb,"ISA Reference",ih,ir,ifmt)
    out=os.path.join(FA,"WYVERN_E2_atmospherics.xlsx"); wb.save(out); return out

# ----------------------------------------------------------------------------- airfoil polars
def build_polars():
    ph,pr = read_csv(os.path.join(CFD,"airfoil_polars.csv"))
    wb=Workbook(); wb.remove(wb.active)
    s=wb.create_sheet("Summary"); title_row(s,"WYVERN-E 2.0 — Fin Airfoil Polars (vortex panel CFD)",4)
    # header for the per-profile summary
    for j,h in enumerate(["Profile","dCl/dα [/deg] (0–6°)","Cl @ 5°","Best L/D (flight)"]):
        c=s.cell(2,j+1,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=Alignment(horizontal="center",wrap_text=True)
    profiles=["NACA0006","NACA0012","double_wedge","flat_plate"]
    block=25                      # rows per profile in the CSV
    for k,prof in enumerate(profiles):
        r0=2+k*block; r1=r0+block-1            # data rows in Polars sheet
        slope_r0=r0; slope_r1=r0+12            # 0..6 deg = first 13 rows
        rr=3+k
        s.cell(rr,1,prof)
        s.cell(rr,2,f"=SLOPE(Polars!D{slope_r0}:D{slope_r1},Polars!B{slope_r0}:B{slope_r1})").number_format="0.0000"
        s.cell(rr,3,f"=INDEX(Polars!D{r0}:D{r1},MATCH(5,Polars!B{r0}:B{r1},0))").number_format="0.000"
        s.cell(rr,4,f"=MAX(Polars!I{r0}:I{r1})").number_format="0.0"
        for c in range(1,5):
            cell=s.cell(rr,c); cell.border=BORDER
            if c==1: cell.font=Font(name="Arial",bold=True,size=10)
            else:    cell.font=BASE; cell.fill=KEYFILL
    s.cell(9,1,"Thin-airfoil benchmark dCl/dα = 2π/rad = 0.1097/deg. Panel method is inviscid: Cl & Cp only;").font=Font(name="Arial",italic=True,size=9,color="808080")
    s.cell(10,1,"viscous Cd is a flat-plate-friction estimate. Real Cd/stall come from the RQ1/RQ2 wind tunnel.").font=Font(name="Arial",italic=True,size=9,color="808080")
    for col,w in zip("ABCD",(14,20,12,16)): s.column_dimensions[col].width=w
    pfmt={"alpha_deg":"0.0","tc":"0.000","Cl_panel":"0.0000","Cl_thin":"0.0000",
          "Cd_tunnel":"0.00000","Cd_flight":"0.00000","LD_tunnel":"0.0","LD_flight":"0.0"}
    data_sheet(wb,"Polars",ph,pr,pfmt)
    out=os.path.join(CFD,"WYVERN_E2_airfoil_polars.xlsx"); wb.save(out); return out

if __name__=="__main__":
    a=build_atmospherics(); print("wrote",os.path.relpath(a,HERE))
    b=build_polars();       print("wrote",os.path.relpath(b,HERE))
